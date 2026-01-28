"""
TekVwarho ProAudit - Financial Report Export Service

Comprehensive service for exporting financial reports in PDF and Excel formats:
- Balance Sheet
- Income Statement (Profit & Loss)
- Trial Balance
- Cash Flow Statement
- General Ledger
- Custom report templates

Nigerian IFRS compliant formatting with company branding support.
"""

import uuid
import io
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional, Dict, Any, Tuple
from enum import Enum as PyEnum

from sqlalchemy import select, func, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession

# PDF Generation (reportlab)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, Image, HRFlowable
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

# Excel Generation (openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.models.accounting import (
    ChartOfAccounts, AccountType, AccountSubType,
    JournalEntry, JournalEntryLine, JournalEntryStatus,
    FiscalYear, FiscalPeriod, AccountBalance
)
from app.models.entity import BusinessEntity


class ReportFormat(str, PyEnum):
    """Export format options"""
    PDF = "pdf"
    EXCEL = "xlsx"
    CSV = "csv"


class ReportType(str, PyEnum):
    """Available report types"""
    BALANCE_SHEET = "balance_sheet"
    INCOME_STATEMENT = "income_statement"
    TRIAL_BALANCE = "trial_balance"
    CASH_FLOW = "cash_flow"
    GENERAL_LEDGER = "general_ledger"
    ACCOUNT_TRANSACTIONS = "account_transactions"


class FinancialReportExportService:
    """Service for generating and exporting financial reports."""
    
    def __init__(self, db: AsyncSession):
        self.db = db
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles for reports."""
        # Title style
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20,
            textColor=colors.HexColor("#1a365d")
        ))
        
        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='ReportSubtitle',
            parent=self.styles['Heading2'],
            fontSize=12,
            alignment=TA_CENTER,
            spaceAfter=10,
            textColor=colors.HexColor("#4a5568")
        ))
        
        # Section header
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading3'],
            fontSize=12,
            fontName='Helvetica-Bold',
            spaceBefore=15,
            spaceAfter=5,
            textColor=colors.HexColor("#2d3748")
        ))
        
        # Account name
        self.styles.add(ParagraphStyle(
            name='AccountName',
            parent=self.styles['Normal'],
            fontSize=9,
            leftIndent=10
        ))
        
        # Footer style
        self.styles.add(ParagraphStyle(
            name='Footer',
            parent=self.styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#718096")
        ))
    
    # =========================================================================
    # MAIN EXPORT METHODS
    # =========================================================================
    
    async def export_balance_sheet(
        self,
        entity_id: uuid.UUID,
        as_of_date: date,
        format: ReportFormat = ReportFormat.PDF,
        comparative_date: Optional[date] = None
    ) -> Tuple[bytes, str]:
        """Export Balance Sheet report."""
        # Get entity details
        entity = await self._get_entity(entity_id)
        
        # Get balance sheet data
        data = await self._get_balance_sheet_data(entity_id, as_of_date, comparative_date)
        
        if format == ReportFormat.PDF:
            return self._generate_balance_sheet_pdf(entity, data, as_of_date, comparative_date)
        elif format == ReportFormat.EXCEL:
            return self._generate_balance_sheet_excel(entity, data, as_of_date, comparative_date)
        else:
            return self._generate_balance_sheet_csv(entity, data, as_of_date)
    
    async def export_income_statement(
        self,
        entity_id: uuid.UUID,
        start_date: date,
        end_date: date,
        format: ReportFormat = ReportFormat.PDF,
        comparative_start: Optional[date] = None,
        comparative_end: Optional[date] = None
    ) -> Tuple[bytes, str]:
        """Export Income Statement (P&L) report."""
        entity = await self._get_entity(entity_id)
        
        # Get income statement data
        data = await self._get_income_statement_data(
            entity_id, start_date, end_date, comparative_start, comparative_end
        )
        
        if format == ReportFormat.PDF:
            return self._generate_income_statement_pdf(entity, data, start_date, end_date)
        elif format == ReportFormat.EXCEL:
            return self._generate_income_statement_excel(entity, data, start_date, end_date)
        else:
            return self._generate_income_statement_csv(entity, data, start_date, end_date)
    
    async def export_trial_balance(
        self,
        entity_id: uuid.UUID,
        as_of_date: date,
        format: ReportFormat = ReportFormat.PDF,
        include_zero_balances: bool = False
    ) -> Tuple[bytes, str]:
        """Export Trial Balance report."""
        entity = await self._get_entity(entity_id)
        
        # Get trial balance data
        data = await self._get_trial_balance_data(entity_id, as_of_date, include_zero_balances)
        
        if format == ReportFormat.PDF:
            return self._generate_trial_balance_pdf(entity, data, as_of_date)
        elif format == ReportFormat.EXCEL:
            return self._generate_trial_balance_excel(entity, data, as_of_date)
        else:
            return self._generate_trial_balance_csv(entity, data, as_of_date)
    
    async def export_general_ledger(
        self,
        entity_id: uuid.UUID,
        start_date: date,
        end_date: date,
        account_id: Optional[uuid.UUID] = None,
        format: ReportFormat = ReportFormat.PDF
    ) -> Tuple[bytes, str]:
        """Export General Ledger report."""
        entity = await self._get_entity(entity_id)
        
        # Get general ledger data
        data = await self._get_general_ledger_data(entity_id, start_date, end_date, account_id)
        
        if format == ReportFormat.PDF:
            return self._generate_general_ledger_pdf(entity, data, start_date, end_date)
        elif format == ReportFormat.EXCEL:
            return self._generate_general_ledger_excel(entity, data, start_date, end_date)
        else:
            return self._generate_general_ledger_csv(entity, data, start_date, end_date)
    
    # =========================================================================
    # DATA RETRIEVAL METHODS
    # =========================================================================
    
    async def _get_entity(self, entity_id: uuid.UUID) -> BusinessEntity:
        """Get entity details."""
        result = await self.db.execute(
            select(BusinessEntity).where(BusinessEntity.id == entity_id)
        )
        entity = result.scalar_one_or_none()
        
        if not entity:
            raise ValueError("Entity not found")
        
        return entity
    
    async def _get_balance_sheet_data(
        self,
        entity_id: uuid.UUID,
        as_of_date: date,
        comparative_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """Get balance sheet data."""
        # Get asset accounts
        assets = await self._get_account_balances_by_type(
            entity_id, AccountType.ASSET, as_of_date
        )
        
        # Get liability accounts
        liabilities = await self._get_account_balances_by_type(
            entity_id, AccountType.LIABILITY, as_of_date
        )
        
        # Get equity accounts
        equity = await self._get_account_balances_by_type(
            entity_id, AccountType.EQUITY, as_of_date
        )
        
        # Calculate retained earnings (current year)
        retained_earnings = await self._calculate_retained_earnings(entity_id, as_of_date)
        
        # Calculate totals
        total_assets = sum(a["balance"] for a in assets)
        total_liabilities = sum(l["balance"] for l in liabilities)
        total_equity = sum(e["balance"] for e in equity) + retained_earnings
        
        data = {
            "assets": {
                "accounts": assets,
                "total": total_assets
            },
            "liabilities": {
                "accounts": liabilities,
                "total": total_liabilities
            },
            "equity": {
                "accounts": equity,
                "retained_earnings": retained_earnings,
                "total": total_equity
            },
            "total_liabilities_equity": total_liabilities + total_equity,
            "is_balanced": abs(total_assets - (total_liabilities + total_equity)) < Decimal("0.01")
        }
        
        # Get comparative data if requested
        if comparative_date:
            comp_assets = await self._get_account_balances_by_type(
                entity_id, AccountType.ASSET, comparative_date
            )
            comp_liabilities = await self._get_account_balances_by_type(
                entity_id, AccountType.LIABILITY, comparative_date
            )
            comp_equity = await self._get_account_balances_by_type(
                entity_id, AccountType.EQUITY, comparative_date
            )
            comp_retained = await self._calculate_retained_earnings(entity_id, comparative_date)
            
            data["comparative"] = {
                "assets": sum(a["balance"] for a in comp_assets),
                "liabilities": sum(l["balance"] for l in comp_liabilities),
                "equity": sum(e["balance"] for e in comp_equity) + comp_retained
            }
        
        return data
    
    async def _get_income_statement_data(
        self,
        entity_id: uuid.UUID,
        start_date: date,
        end_date: date,
        comparative_start: Optional[date] = None,
        comparative_end: Optional[date] = None
    ) -> Dict[str, Any]:
        """Get income statement data."""
        # Get revenue accounts
        revenue = await self._get_period_activity_by_type(
            entity_id, AccountType.REVENUE, start_date, end_date
        )
        
        # Get expense accounts
        expenses = await self._get_period_activity_by_type(
            entity_id, AccountType.EXPENSE, start_date, end_date
        )
        
        total_revenue = sum(r["balance"] for r in revenue)
        total_expenses = sum(e["balance"] for e in expenses)
        net_income = total_revenue - total_expenses
        
        data = {
            "revenue": {
                "accounts": revenue,
                "total": total_revenue
            },
            "expenses": {
                "accounts": expenses,
                "total": total_expenses
            },
            "net_income": net_income,
            "profit_margin": (net_income / total_revenue * 100) if total_revenue else Decimal("0")
        }
        
        # Get comparative data if requested
        if comparative_start and comparative_end:
            comp_revenue = await self._get_period_activity_by_type(
                entity_id, AccountType.REVENUE, comparative_start, comparative_end
            )
            comp_expenses = await self._get_period_activity_by_type(
                entity_id, AccountType.EXPENSE, comparative_start, comparative_end
            )
            
            comp_total_rev = sum(r["balance"] for r in comp_revenue)
            comp_total_exp = sum(e["balance"] for e in comp_expenses)
            
            data["comparative"] = {
                "revenue": comp_total_rev,
                "expenses": comp_total_exp,
                "net_income": comp_total_rev - comp_total_exp
            }
        
        return data
    
    async def _get_trial_balance_data(
        self,
        entity_id: uuid.UUID,
        as_of_date: date,
        include_zero: bool = False
    ) -> Dict[str, Any]:
        """Get trial balance data."""
        # Get all accounts with balances
        result = await self.db.execute(
            select(ChartOfAccounts).where(
                and_(
                    ChartOfAccounts.entity_id == entity_id,
                    ChartOfAccounts.is_active == True,
                    ChartOfAccounts.is_header == False
                )
            ).order_by(ChartOfAccounts.account_code)
        )
        accounts = list(result.scalars().all())
        
        trial_balance = []
        total_debits = Decimal("0")
        total_credits = Decimal("0")
        
        for account in accounts:
            # Get account balance
            balance_result = await self.db.execute(
                select(
                    func.sum(AccountBalance.debit_balance).label("debit"),
                    func.sum(AccountBalance.credit_balance).label("credit")
                ).where(
                    and_(
                        AccountBalance.account_id == account.id,
                        AccountBalance.period_end_date <= as_of_date
                    )
                )
            )
            row = balance_result.one_or_none()
            
            debit = row[0] or Decimal("0") if row else Decimal("0")
            credit = row[1] or Decimal("0") if row else Decimal("0")
            
            if not include_zero and debit == Decimal("0") and credit == Decimal("0"):
                continue
            
            trial_balance.append({
                "account_code": account.account_code,
                "account_name": account.account_name,
                "account_type": account.account_type.value,
                "debit": debit,
                "credit": credit
            })
            
            total_debits += debit
            total_credits += credit
        
        return {
            "accounts": trial_balance,
            "total_debits": total_debits,
            "total_credits": total_credits,
            "is_balanced": abs(total_debits - total_credits) < Decimal("0.01"),
            "difference": abs(total_debits - total_credits)
        }
    
    async def _get_general_ledger_data(
        self,
        entity_id: uuid.UUID,
        start_date: date,
        end_date: date,
        account_id: Optional[uuid.UUID] = None
    ) -> Dict[str, Any]:
        """Get general ledger data."""
        # Build account query
        account_query = select(ChartOfAccounts).where(
            and_(
                ChartOfAccounts.entity_id == entity_id,
                ChartOfAccounts.is_active == True,
                ChartOfAccounts.is_header == False
            )
        )
        
        if account_id:
            account_query = account_query.where(ChartOfAccounts.id == account_id)
        
        account_query = account_query.order_by(ChartOfAccounts.account_code)
        
        result = await self.db.execute(account_query)
        accounts = list(result.scalars().all())
        
        ledger_data = []
        
        for account in accounts:
            # Get transactions for this account
            transactions = await self._get_account_transactions(
                account.id, start_date, end_date
            )
            
            if transactions:
                ledger_data.append({
                    "account_code": account.account_code,
                    "account_name": account.account_name,
                    "account_type": account.account_type.value,
                    "transactions": transactions,
                    "opening_balance": await self._get_opening_balance(account.id, start_date),
                    "closing_balance": await self._get_closing_balance(account.id, end_date)
                })
        
        return {
            "accounts": ledger_data,
            "period": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat()
            }
        }
    
    async def _get_account_balances_by_type(
        self,
        entity_id: uuid.UUID,
        account_type: AccountType,
        as_of_date: date
    ) -> List[Dict[str, Any]]:
        """Get account balances for a specific account type."""
        result = await self.db.execute(
            select(ChartOfAccounts).where(
                and_(
                    ChartOfAccounts.entity_id == entity_id,
                    ChartOfAccounts.account_type == account_type,
                    ChartOfAccounts.is_active == True,
                    ChartOfAccounts.is_header == False
                )
            ).order_by(ChartOfAccounts.account_code)
        )
        accounts = list(result.scalars().all())
        
        balances = []
        for account in accounts:
            balance = await self._get_closing_balance(account.id, as_of_date)
            if balance != Decimal("0"):
                balances.append({
                    "code": account.account_code,
                    "name": account.account_name,
                    "sub_type": account.account_sub_type.value if account.account_sub_type else None,
                    "balance": balance
                })
        
        return balances
    
    async def _get_period_activity_by_type(
        self,
        entity_id: uuid.UUID,
        account_type: AccountType,
        start_date: date,
        end_date: date
    ) -> List[Dict[str, Any]]:
        """Get account activity for a period for a specific account type."""
        result = await self.db.execute(
            select(ChartOfAccounts).where(
                and_(
                    ChartOfAccounts.entity_id == entity_id,
                    ChartOfAccounts.account_type == account_type,
                    ChartOfAccounts.is_active == True,
                    ChartOfAccounts.is_header == False
                )
            ).order_by(ChartOfAccounts.account_code)
        )
        accounts = list(result.scalars().all())
        
        activities = []
        for account in accounts:
            # Calculate activity for the period
            activity_result = await self.db.execute(
                select(
                    func.sum(JournalEntryLine.debit_amount).label("debit"),
                    func.sum(JournalEntryLine.credit_amount).label("credit")
                ).select_from(JournalEntryLine).join(
                    JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
                ).where(
                    and_(
                        JournalEntryLine.account_id == account.id,
                        JournalEntry.entry_date >= start_date,
                        JournalEntry.entry_date <= end_date,
                        JournalEntry.status == JournalEntryStatus.POSTED
                    )
                )
            )
            row = activity_result.one_or_none()
            
            debit = row[0] or Decimal("0") if row else Decimal("0")
            credit = row[1] or Decimal("0") if row else Decimal("0")
            
            # For revenue, credit is positive; for expense, debit is positive
            if account_type == AccountType.REVENUE:
                balance = credit - debit
            else:
                balance = debit - credit
            
            if balance != Decimal("0"):
                activities.append({
                    "code": account.account_code,
                    "name": account.account_name,
                    "sub_type": account.account_sub_type.value if account.account_sub_type else None,
                    "balance": balance
                })
        
        return activities
    
    async def _calculate_retained_earnings(
        self,
        entity_id: uuid.UUID,
        as_of_date: date
    ) -> Decimal:
        """Calculate retained earnings (accumulated profit/loss)."""
        # Get total revenue
        revenue_result = await self.db.execute(
            select(func.sum(JournalEntryLine.credit_amount - JournalEntryLine.debit_amount)).select_from(
                JournalEntryLine
            ).join(
                JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
            ).join(
                ChartOfAccounts, JournalEntryLine.account_id == ChartOfAccounts.id
            ).where(
                and_(
                    JournalEntry.entity_id == entity_id,
                    JournalEntry.entry_date <= as_of_date,
                    JournalEntry.status == JournalEntryStatus.POSTED,
                    ChartOfAccounts.account_type == AccountType.REVENUE
                )
            )
        )
        total_revenue = revenue_result.scalar() or Decimal("0")
        
        # Get total expenses
        expense_result = await self.db.execute(
            select(func.sum(JournalEntryLine.debit_amount - JournalEntryLine.credit_amount)).select_from(
                JournalEntryLine
            ).join(
                JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
            ).join(
                ChartOfAccounts, JournalEntryLine.account_id == ChartOfAccounts.id
            ).where(
                and_(
                    JournalEntry.entity_id == entity_id,
                    JournalEntry.entry_date <= as_of_date,
                    JournalEntry.status == JournalEntryStatus.POSTED,
                    ChartOfAccounts.account_type == AccountType.EXPENSE
                )
            )
        )
        total_expenses = expense_result.scalar() or Decimal("0")
        
        return total_revenue - total_expenses
    
    async def _get_account_transactions(
        self,
        account_id: uuid.UUID,
        start_date: date,
        end_date: date
    ) -> List[Dict[str, Any]]:
        """Get transactions for an account within a date range."""
        result = await self.db.execute(
            select(JournalEntryLine, JournalEntry).join(
                JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
            ).where(
                and_(
                    JournalEntryLine.account_id == account_id,
                    JournalEntry.entry_date >= start_date,
                    JournalEntry.entry_date <= end_date,
                    JournalEntry.status == JournalEntryStatus.POSTED
                )
            ).order_by(JournalEntry.entry_date, JournalEntry.entry_number)
        )
        
        transactions = []
        for line, entry in result.all():
            transactions.append({
                "date": entry.entry_date.isoformat(),
                "entry_number": entry.entry_number,
                "description": line.description or entry.description,
                "debit": float(line.debit_amount) if line.debit_amount else 0,
                "credit": float(line.credit_amount) if line.credit_amount else 0
            })
        
        return transactions
    
    async def _get_opening_balance(
        self,
        account_id: uuid.UUID,
        as_of_date: date
    ) -> Decimal:
        """Get opening balance for an account."""
        result = await self.db.execute(
            select(
                func.sum(JournalEntryLine.debit_amount - JournalEntryLine.credit_amount)
            ).select_from(JournalEntryLine).join(
                JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
            ).where(
                and_(
                    JournalEntryLine.account_id == account_id,
                    JournalEntry.entry_date < as_of_date,
                    JournalEntry.status == JournalEntryStatus.POSTED
                )
            )
        )
        return result.scalar() or Decimal("0")
    
    async def _get_closing_balance(
        self,
        account_id: uuid.UUID,
        as_of_date: date
    ) -> Decimal:
        """Get closing balance for an account."""
        result = await self.db.execute(
            select(
                func.sum(JournalEntryLine.debit_amount - JournalEntryLine.credit_amount)
            ).select_from(JournalEntryLine).join(
                JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id
            ).where(
                and_(
                    JournalEntryLine.account_id == account_id,
                    JournalEntry.entry_date <= as_of_date,
                    JournalEntry.status == JournalEntryStatus.POSTED
                )
            )
        )
        return result.scalar() or Decimal("0")
    
    # =========================================================================
    # PDF GENERATION METHODS
    # =========================================================================
    
    def _generate_balance_sheet_pdf(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        as_of_date: date,
        comparative_date: Optional[date] = None
    ) -> Tuple[bytes, str]:
        """Generate Balance Sheet PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        
        # Header
        elements.append(Paragraph(entity.name, self.styles['ReportTitle']))
        elements.append(Paragraph(
            f"Statement of Financial Position (Balance Sheet)",
            self.styles['ReportSubtitle']
        ))
        elements.append(Paragraph(
            f"As of {as_of_date.strftime('%B %d, %Y')}",
            self.styles['ReportSubtitle']
        ))
        elements.append(Spacer(1, 20))
        
        # Currency note
        elements.append(Paragraph(
            "All amounts in Nigerian Naira (₦)",
            self.styles['Footer']
        ))
        elements.append(Spacer(1, 10))
        
        # Assets Section
        elements.append(Paragraph("ASSETS", self.styles['SectionHeader']))
        
        asset_data = [["Account", "Amount (₦)"]]
        for acc in data["assets"]["accounts"]:
            asset_data.append([
                f"  {acc['code']} - {acc['name']}",
                self._format_currency(acc['balance'])
            ])
        asset_data.append(["Total Assets", self._format_currency(data["assets"]["total"])])
        
        asset_table = self._create_table(asset_data, highlight_last_row=True)
        elements.append(asset_table)
        elements.append(Spacer(1, 15))
        
        # Liabilities Section
        elements.append(Paragraph("LIABILITIES", self.styles['SectionHeader']))
        
        liability_data = [["Account", "Amount (₦)"]]
        for acc in data["liabilities"]["accounts"]:
            liability_data.append([
                f"  {acc['code']} - {acc['name']}",
                self._format_currency(acc['balance'])
            ])
        liability_data.append(["Total Liabilities", self._format_currency(data["liabilities"]["total"])])
        
        liability_table = self._create_table(liability_data, highlight_last_row=True)
        elements.append(liability_table)
        elements.append(Spacer(1, 15))
        
        # Equity Section
        elements.append(Paragraph("EQUITY", self.styles['SectionHeader']))
        
        equity_data = [["Account", "Amount (₦)"]]
        for acc in data["equity"]["accounts"]:
            equity_data.append([
                f"  {acc['code']} - {acc['name']}",
                self._format_currency(acc['balance'])
            ])
        if data["equity"]["retained_earnings"] != Decimal("0"):
            equity_data.append([
                "  Current Year Earnings",
                self._format_currency(data["equity"]["retained_earnings"])
            ])
        equity_data.append(["Total Equity", self._format_currency(data["equity"]["total"])])
        
        equity_table = self._create_table(equity_data, highlight_last_row=True)
        elements.append(equity_table)
        elements.append(Spacer(1, 20))
        
        # Summary
        elements.append(HRFlowable(width="100%", color=colors.black))
        summary_data = [
            ["Total Liabilities & Equity", self._format_currency(data["total_liabilities_equity"])]
        ]
        summary_table = self._create_table(summary_data, is_summary=True)
        elements.append(summary_table)
        
        # Balance check
        if data["is_balanced"]:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(
                "✓ Balance Sheet is balanced",
                self.styles['Footer']
            ))
        
        # Footer
        elements.append(Spacer(1, 30))
        elements.append(HRFlowable(width="100%", color=colors.gray))
        elements.append(Paragraph(
            f"Generated by TekVwarho ProAudit on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            self.styles['Footer']
        ))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"balance_sheet_{entity.name}_{as_of_date.strftime('%Y%m%d')}.pdf"
        return buffer.read(), filename
    
    def _generate_income_statement_pdf(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        start_date: date,
        end_date: date
    ) -> Tuple[bytes, str]:
        """Generate Income Statement PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        
        # Header
        elements.append(Paragraph(entity.name, self.styles['ReportTitle']))
        elements.append(Paragraph(
            "Statement of Profit or Loss (Income Statement)",
            self.styles['ReportSubtitle']
        ))
        elements.append(Paragraph(
            f"For the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
            self.styles['ReportSubtitle']
        ))
        elements.append(Spacer(1, 20))
        
        # Currency note
        elements.append(Paragraph(
            "All amounts in Nigerian Naira (₦)",
            self.styles['Footer']
        ))
        elements.append(Spacer(1, 10))
        
        # Revenue Section
        elements.append(Paragraph("REVENUE", self.styles['SectionHeader']))
        
        revenue_data = [["Account", "Amount (₦)"]]
        for acc in data["revenue"]["accounts"]:
            revenue_data.append([
                f"  {acc['code']} - {acc['name']}",
                self._format_currency(acc['balance'])
            ])
        revenue_data.append(["Total Revenue", self._format_currency(data["revenue"]["total"])])
        
        revenue_table = self._create_table(revenue_data, highlight_last_row=True)
        elements.append(revenue_table)
        elements.append(Spacer(1, 15))
        
        # Expenses Section
        elements.append(Paragraph("EXPENSES", self.styles['SectionHeader']))
        
        expense_data = [["Account", "Amount (₦)"]]
        for acc in data["expenses"]["accounts"]:
            expense_data.append([
                f"  {acc['code']} - {acc['name']}",
                self._format_currency(acc['balance'])
            ])
        expense_data.append(["Total Expenses", self._format_currency(data["expenses"]["total"])])
        
        expense_table = self._create_table(expense_data, highlight_last_row=True)
        elements.append(expense_table)
        elements.append(Spacer(1, 20))
        
        # Net Income Summary
        elements.append(HRFlowable(width="100%", color=colors.black))
        
        net_income = data["net_income"]
        profit_loss_label = "Net Profit" if net_income >= 0 else "Net Loss"
        
        summary_data = [
            [profit_loss_label, self._format_currency(abs(net_income))]
        ]
        
        summary_table = self._create_table(summary_data, is_summary=True, is_profit=net_income >= 0)
        elements.append(summary_table)
        
        # Profit margin
        if data["profit_margin"] != Decimal("0"):
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(
                f"Profit Margin: {data['profit_margin']:.1f}%",
                self.styles['Footer']
            ))
        
        # Footer
        elements.append(Spacer(1, 30))
        elements.append(HRFlowable(width="100%", color=colors.gray))
        elements.append(Paragraph(
            f"Generated by TekVwarho ProAudit on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            self.styles['Footer']
        ))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"income_statement_{entity.name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
        return buffer.read(), filename
    
    def _generate_trial_balance_pdf(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        as_of_date: date
    ) -> Tuple[bytes, str]:
        """Generate Trial Balance PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        
        # Header
        elements.append(Paragraph(entity.name, self.styles['ReportTitle']))
        elements.append(Paragraph("Trial Balance", self.styles['ReportSubtitle']))
        elements.append(Paragraph(
            f"As of {as_of_date.strftime('%B %d, %Y')}",
            self.styles['ReportSubtitle']
        ))
        elements.append(Spacer(1, 20))
        
        # Currency note
        elements.append(Paragraph(
            "All amounts in Nigerian Naira (₦)",
            self.styles['Footer']
        ))
        elements.append(Spacer(1, 10))
        
        # Trial Balance Table
        tb_data = [["Account Code", "Account Name", "Debit (₦)", "Credit (₦)"]]
        
        for acc in data["accounts"]:
            tb_data.append([
                acc["account_code"],
                acc["account_name"],
                self._format_currency(acc["debit"]) if acc["debit"] > 0 else "",
                self._format_currency(acc["credit"]) if acc["credit"] > 0 else ""
            ])
        
        # Totals row
        tb_data.append([
            "", "TOTALS",
            self._format_currency(data["total_debits"]),
            self._format_currency(data["total_credits"])
        ])
        
        tb_table = Table(tb_data, colWidths=[1.2*inch, 3.5*inch, 1.3*inch, 1.3*inch])
        tb_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2d3748")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#e2e8f0")),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (2, 1), (3, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ]))
        elements.append(tb_table)
        
        # Balance check
        elements.append(Spacer(1, 15))
        if data["is_balanced"]:
            elements.append(Paragraph(
                "✓ Trial Balance is balanced (Debits = Credits)",
                self.styles['Footer']
            ))
        else:
            elements.append(Paragraph(
                f"⚠ Trial Balance is out of balance by ₦{self._format_currency(data['difference'])}",
                ParagraphStyle(
                    'Warning',
                    parent=self.styles['Footer'],
                    textColor=colors.red
                )
            ))
        
        # Footer
        elements.append(Spacer(1, 30))
        elements.append(HRFlowable(width="100%", color=colors.gray))
        elements.append(Paragraph(
            f"Generated by TekVwarho ProAudit on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            self.styles['Footer']
        ))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"trial_balance_{entity.name}_{as_of_date.strftime('%Y%m%d')}.pdf"
        return buffer.read(), filename
    
    def _generate_general_ledger_pdf(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        start_date: date,
        end_date: date
    ) -> Tuple[bytes, str]:
        """Generate General Ledger PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        
        # Header
        elements.append(Paragraph(entity.name, self.styles['ReportTitle']))
        elements.append(Paragraph("General Ledger", self.styles['ReportSubtitle']))
        elements.append(Paragraph(
            f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
            self.styles['ReportSubtitle']
        ))
        elements.append(Spacer(1, 20))
        
        # Each account
        for account in data["accounts"]:
            # Account header
            elements.append(Paragraph(
                f"{account['account_code']} - {account['account_name']}",
                self.styles['SectionHeader']
            ))
            
            # Opening balance
            elements.append(Paragraph(
                f"Opening Balance: ₦{self._format_currency(account['opening_balance'])}",
                self.styles['Normal']
            ))
            elements.append(Spacer(1, 5))
            
            # Transactions table
            txn_data = [["Date", "Entry #", "Description", "Debit", "Credit"]]
            
            for txn in account["transactions"]:
                txn_data.append([
                    txn["date"],
                    txn["entry_number"],
                    txn["description"][:40] if txn["description"] else "",
                    self._format_currency(Decimal(str(txn["debit"]))) if txn["debit"] > 0 else "",
                    self._format_currency(Decimal(str(txn["credit"]))) if txn["credit"] > 0 else ""
                ])
            
            txn_table = Table(txn_data, colWidths=[0.8*inch, 0.9*inch, 3*inch, 0.9*inch, 0.9*inch])
            txn_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(txn_table)
            
            # Closing balance
            elements.append(Paragraph(
                f"Closing Balance: ₦{self._format_currency(account['closing_balance'])}",
                self.styles['Normal']
            ))
            elements.append(Spacer(1, 20))
        
        # Footer
        elements.append(HRFlowable(width="100%", color=colors.gray))
        elements.append(Paragraph(
            f"Generated by TekVwarho ProAudit on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            self.styles['Footer']
        ))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"general_ledger_{entity.name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
        return buffer.read(), filename
    
    # =========================================================================
    # EXCEL GENERATION METHODS
    # =========================================================================
    
    def _generate_balance_sheet_excel(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        as_of_date: date,
        comparative_date: Optional[date] = None
    ) -> Tuple[bytes, str]:
        """Generate Balance Sheet Excel."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed. Install it with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        header_font_white = Font(bold=True, size=10, color="FFFFFF")
        currency_format = '#,##0.00'
        
        # Title
        ws['A1'] = entity.name
        ws['A1'].font = title_font
        ws['A2'] = "Statement of Financial Position (Balance Sheet)"
        ws['A3'] = f"As of {as_of_date.strftime('%B %d, %Y')}"
        
        row = 5
        
        # Assets
        ws[f'A{row}'] = "ASSETS"
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = "Amount (₦)"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font_white
        ws[f'B{row}'].font = header_font_white
        row += 1
        
        for acc in data["assets"]["accounts"]:
            ws[f'A{row}'] = f"  {acc['code']} - {acc['name']}"
            ws[f'B{row}'] = float(acc['balance'])
            ws[f'B{row}'].number_format = currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Assets"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = float(data["assets"]["total"])
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].font = header_font
        row += 2
        
        # Liabilities
        ws[f'A{row}'] = "LIABILITIES"
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = "Amount (₦)"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font_white
        ws[f'B{row}'].font = header_font_white
        row += 1
        
        for acc in data["liabilities"]["accounts"]:
            ws[f'A{row}'] = f"  {acc['code']} - {acc['name']}"
            ws[f'B{row}'] = float(acc['balance'])
            ws[f'B{row}'].number_format = currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Liabilities"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = float(data["liabilities"]["total"])
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].font = header_font
        row += 2
        
        # Equity
        ws[f'A{row}'] = "EQUITY"
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = "Amount (₦)"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font_white
        ws[f'B{row}'].font = header_font_white
        row += 1
        
        for acc in data["equity"]["accounts"]:
            ws[f'A{row}'] = f"  {acc['code']} - {acc['name']}"
            ws[f'B{row}'] = float(acc['balance'])
            ws[f'B{row}'].number_format = currency_format
            row += 1
        
        if data["equity"]["retained_earnings"] != Decimal("0"):
            ws[f'A{row}'] = "  Current Year Earnings"
            ws[f'B{row}'] = float(data["equity"]["retained_earnings"])
            ws[f'B{row}'].number_format = currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Equity"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = float(data["equity"]["total"])
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].font = header_font
        row += 2
        
        # Total
        ws[f'A{row}'] = "Total Liabilities & Equity"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'B{row}'] = float(data["total_liabilities_equity"])
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].font = Font(bold=True, size=11)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 20
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"balance_sheet_{entity.name}_{as_of_date.strftime('%Y%m%d')}.xlsx"
        return buffer.read(), filename
    
    def _generate_income_statement_excel(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        start_date: date,
        end_date: date
    ) -> Tuple[bytes, str]:
        """Generate Income Statement Excel."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        header_font_white = Font(bold=True, size=10, color="FFFFFF")
        currency_format = '#,##0.00'
        
        # Title
        ws['A1'] = entity.name
        ws['A1'].font = title_font
        ws['A2'] = "Statement of Profit or Loss (Income Statement)"
        ws['A3'] = f"For the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        
        row = 5
        
        # Revenue
        ws[f'A{row}'] = "REVENUE"
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = "Amount (₦)"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font_white
        ws[f'B{row}'].font = header_font_white
        row += 1
        
        for acc in data["revenue"]["accounts"]:
            ws[f'A{row}'] = f"  {acc['code']} - {acc['name']}"
            ws[f'B{row}'] = float(acc['balance'])
            ws[f'B{row}'].number_format = currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Revenue"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = float(data["revenue"]["total"])
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].font = header_font
        row += 2
        
        # Expenses
        ws[f'A{row}'] = "EXPENSES"
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = "Amount (₦)"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font_white
        ws[f'B{row}'].font = header_font_white
        row += 1
        
        for acc in data["expenses"]["accounts"]:
            ws[f'A{row}'] = f"  {acc['code']} - {acc['name']}"
            ws[f'B{row}'] = float(acc['balance'])
            ws[f'B{row}'].number_format = currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Expenses"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = float(data["expenses"]["total"])
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].font = header_font
        row += 2
        
        # Net Income
        net_income = data["net_income"]
        label = "Net Profit" if net_income >= 0 else "Net Loss"
        
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = float(abs(net_income))
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].font = Font(bold=True, size=12, color="008000" if net_income >= 0 else "FF0000")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 20
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"income_statement_{entity.name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        return buffer.read(), filename
    
    def _generate_trial_balance_excel(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        as_of_date: date
    ) -> Tuple[bytes, str]:
        """Generate Trial Balance Excel."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Trial Balance"
        
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        header_font_white = Font(bold=True, size=10, color="FFFFFF")
        currency_format = '#,##0.00'
        
        # Title
        ws['A1'] = entity.name
        ws['A1'].font = title_font
        ws['A2'] = "Trial Balance"
        ws['A3'] = f"As of {as_of_date.strftime('%B %d, %Y')}"
        
        # Headers
        headers = ["Account Code", "Account Name", "Debit (₦)", "Credit (₦)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font_white
        
        # Data
        row = 6
        for acc in data["accounts"]:
            ws.cell(row=row, column=1, value=acc["account_code"])
            ws.cell(row=row, column=2, value=acc["account_name"])
            if acc["debit"] > 0:
                cell = ws.cell(row=row, column=3, value=float(acc["debit"]))
                cell.number_format = currency_format
            if acc["credit"] > 0:
                cell = ws.cell(row=row, column=4, value=float(acc["credit"]))
                cell.number_format = currency_format
            row += 1
        
        # Totals
        ws.cell(row=row, column=2, value="TOTALS").font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(data["total_debits"])).number_format = currency_format
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(data["total_credits"])).number_format = currency_format
        ws.cell(row=row, column=4).font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"trial_balance_{entity.name}_{as_of_date.strftime('%Y%m%d')}.xlsx"
        return buffer.read(), filename
    
    def _generate_general_ledger_excel(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        start_date: date,
        end_date: date
    ) -> Tuple[bytes, str]:
        """Generate General Ledger Excel."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")
        
        wb = Workbook()
        
        for i, account in enumerate(data["accounts"]):
            # Create sheet for each account (max 31 chars)
            sheet_name = f"{account['account_code'][:10]}-{account['account_name'][:18]}"
            sheet_name = sheet_name[:31]  # Excel limit
            
            if i == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(sheet_name)
            
            title_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            currency_format = '#,##0.00'
            
            # Account header
            ws['A1'] = f"{account['account_code']} - {account['account_name']}"
            ws['A1'].font = title_font
            ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            ws['A3'] = f"Opening Balance: {float(account['opening_balance']):,.2f}"
            
            # Transaction headers
            headers = ["Date", "Entry #", "Description", "Debit", "Credit", "Balance"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True)
            
            # Transactions
            row = 6
            running_balance = account['opening_balance']
            
            for txn in account["transactions"]:
                ws.cell(row=row, column=1, value=txn["date"])
                ws.cell(row=row, column=2, value=txn["entry_number"])
                ws.cell(row=row, column=3, value=txn["description"] or "")
                
                if txn["debit"] > 0:
                    ws.cell(row=row, column=4, value=txn["debit"]).number_format = currency_format
                    running_balance += Decimal(str(txn["debit"]))
                if txn["credit"] > 0:
                    ws.cell(row=row, column=5, value=txn["credit"]).number_format = currency_format
                    running_balance -= Decimal(str(txn["credit"]))
                
                ws.cell(row=row, column=6, value=float(running_balance)).number_format = currency_format
                row += 1
            
            # Closing balance
            ws.cell(row=row, column=3, value="Closing Balance").font = Font(bold=True)
            ws.cell(row=row, column=6, value=float(account['closing_balance'])).number_format = currency_format
            ws.cell(row=row, column=6).font = Font(bold=True)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"general_ledger_{entity.name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        return buffer.read(), filename
    
    # =========================================================================
    # CSV GENERATION METHODS
    # =========================================================================
    
    def _generate_balance_sheet_csv(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        as_of_date: date
    ) -> Tuple[bytes, str]:
        """Generate Balance Sheet CSV."""
        import csv
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        writer.writerow([entity.name])
        writer.writerow(["Balance Sheet"])
        writer.writerow([f"As of {as_of_date.strftime('%Y-%m-%d')}"])
        writer.writerow([])
        
        writer.writerow(["ASSETS"])
        writer.writerow(["Account Code", "Account Name", "Amount"])
        for acc in data["assets"]["accounts"]:
            writer.writerow([acc["code"], acc["name"], float(acc["balance"])])
        writer.writerow(["", "Total Assets", float(data["assets"]["total"])])
        writer.writerow([])
        
        writer.writerow(["LIABILITIES"])
        writer.writerow(["Account Code", "Account Name", "Amount"])
        for acc in data["liabilities"]["accounts"]:
            writer.writerow([acc["code"], acc["name"], float(acc["balance"])])
        writer.writerow(["", "Total Liabilities", float(data["liabilities"]["total"])])
        writer.writerow([])
        
        writer.writerow(["EQUITY"])
        writer.writerow(["Account Code", "Account Name", "Amount"])
        for acc in data["equity"]["accounts"]:
            writer.writerow([acc["code"], acc["name"], float(acc["balance"])])
        writer.writerow(["", "Retained Earnings", float(data["equity"]["retained_earnings"])])
        writer.writerow(["", "Total Equity", float(data["equity"]["total"])])
        
        content = buffer.getvalue().encode('utf-8')
        filename = f"balance_sheet_{entity.name}_{as_of_date.strftime('%Y%m%d')}.csv"
        return content, filename
    
    def _generate_income_statement_csv(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        start_date: date,
        end_date: date
    ) -> Tuple[bytes, str]:
        """Generate Income Statement CSV."""
        import csv
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        writer.writerow([entity.name])
        writer.writerow(["Income Statement"])
        writer.writerow([f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"])
        writer.writerow([])
        
        writer.writerow(["REVENUE"])
        writer.writerow(["Account Code", "Account Name", "Amount"])
        for acc in data["revenue"]["accounts"]:
            writer.writerow([acc["code"], acc["name"], float(acc["balance"])])
        writer.writerow(["", "Total Revenue", float(data["revenue"]["total"])])
        writer.writerow([])
        
        writer.writerow(["EXPENSES"])
        writer.writerow(["Account Code", "Account Name", "Amount"])
        for acc in data["expenses"]["accounts"]:
            writer.writerow([acc["code"], acc["name"], float(acc["balance"])])
        writer.writerow(["", "Total Expenses", float(data["expenses"]["total"])])
        writer.writerow([])
        
        writer.writerow(["NET INCOME", "", float(data["net_income"])])
        
        content = buffer.getvalue().encode('utf-8')
        filename = f"income_statement_{entity.name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
        return content, filename
    
    def _generate_trial_balance_csv(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        as_of_date: date
    ) -> Tuple[bytes, str]:
        """Generate Trial Balance CSV."""
        import csv
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        writer.writerow([entity.name])
        writer.writerow(["Trial Balance"])
        writer.writerow([f"As of {as_of_date.strftime('%Y-%m-%d')}"])
        writer.writerow([])
        
        writer.writerow(["Account Code", "Account Name", "Debit", "Credit"])
        for acc in data["accounts"]:
            writer.writerow([
                acc["account_code"],
                acc["account_name"],
                float(acc["debit"]) if acc["debit"] > 0 else "",
                float(acc["credit"]) if acc["credit"] > 0 else ""
            ])
        
        writer.writerow(["", "TOTALS", float(data["total_debits"]), float(data["total_credits"])])
        
        content = buffer.getvalue().encode('utf-8')
        filename = f"trial_balance_{entity.name}_{as_of_date.strftime('%Y%m%d')}.csv"
        return content, filename
    
    def _generate_general_ledger_csv(
        self,
        entity: BusinessEntity,
        data: Dict[str, Any],
        start_date: date,
        end_date: date
    ) -> Tuple[bytes, str]:
        """Generate General Ledger CSV."""
        import csv
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        writer.writerow([entity.name])
        writer.writerow(["General Ledger"])
        writer.writerow([f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"])
        writer.writerow([])
        
        for account in data["accounts"]:
            writer.writerow([f"{account['account_code']} - {account['account_name']}"])
            writer.writerow([f"Opening Balance: {float(account['opening_balance'])}"])
            writer.writerow(["Date", "Entry #", "Description", "Debit", "Credit"])
            
            for txn in account["transactions"]:
                writer.writerow([
                    txn["date"],
                    txn["entry_number"],
                    txn["description"] or "",
                    txn["debit"] if txn["debit"] > 0 else "",
                    txn["credit"] if txn["credit"] > 0 else ""
                ])
            
            writer.writerow([f"Closing Balance: {float(account['closing_balance'])}"])
            writer.writerow([])
        
        content = buffer.getvalue().encode('utf-8')
        filename = f"general_ledger_{entity.name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
        return content, filename
    
    # =========================================================================
    # HELPER METHODS
    # =========================================================================
    
    def _format_currency(self, amount: Decimal) -> str:
        """Format amount as currency string."""
        return f"{amount:,.2f}"
    
    def _create_table(
        self,
        data: List[List[str]],
        highlight_last_row: bool = False,
        is_summary: bool = False,
        is_profit: bool = True
    ) -> Table:
        """Create a styled table."""
        table = Table(data, colWidths=[4.5*inch, 2*inch])
        
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#2d3748")),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor("#cbd5e0")),
        ]
        
        if highlight_last_row:
            style_commands.extend([
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor("#2d3748")),
            ])
        
        if is_summary:
            bg_color = colors.HexColor("#48bb78") if is_profit else colors.HexColor("#fc8181")
            style_commands.extend([
                ('BACKGROUND', (0, 0), (-1, -1), bg_color),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
            ])
        
        table.setStyle(TableStyle(style_commands))
        return table

    # =========================================================================
    # CONSOLIDATED FINANCIAL REPORT EXPORTS
    # =========================================================================
    
    async def export_consolidated_balance_sheet(
        self,
        group_id: uuid.UUID,
        as_of_date: date,
        format: ReportFormat = ReportFormat.PDF,
        include_minority_interest: bool = True,
    ) -> Tuple[bytes, str]:
        """
        Export consolidated balance sheet for an entity group.
        
        Args:
            group_id: ID of the entity group
            as_of_date: Balance sheet date
            format: Export format (PDF, XLSX, CSV)
            include_minority_interest: Include non-controlling interest details
            
        Returns:
            Tuple of (file content bytes, filename)
        """
        from app.services.consolidation_service import ConsolidationService
        
        consol_service = ConsolidationService(self.db)
        data = await consol_service.get_consolidated_balance_sheet(
            group_id=group_id,
            as_of_date=as_of_date
        )
        
        # Get group details for header
        group = await consol_service.get_entity_group(group_id)
        group_name = group.name if group else "Consolidated Group"
        
        if format == ReportFormat.PDF:
            return await self._generate_consolidated_balance_sheet_pdf(
                data, group_name, as_of_date, include_minority_interest
            )
        elif format == ReportFormat.EXCEL:
            return await self._generate_consolidated_balance_sheet_excel(
                data, group_name, as_of_date, include_minority_interest
            )
        else:  # CSV
            return await self._generate_consolidated_balance_sheet_csv(
                data, group_name, as_of_date
            )
    
    async def _generate_consolidated_balance_sheet_pdf(
        self,
        data: Dict[str, Any],
        group_name: str,
        as_of_date: date,
        include_minority_interest: bool
    ) -> Tuple[bytes, str]:
        """Generate consolidated balance sheet PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Title
        elements.append(Paragraph(group_name, self.styles['ReportTitle']))
        elements.append(Paragraph("Consolidated Balance Sheet", self.styles['ReportSubtitle']))
        elements.append(Paragraph(f"As of {as_of_date.strftime('%B %d, %Y')}", self.styles['ReportSubtitle']))
        elements.append(Spacer(1, 20))
        
        # Assets section
        elements.append(Paragraph("ASSETS", self.styles['SectionHeader']))
        
        asset_data = [["Account", "Amount"]]
        for asset in data.get("assets", []):
            asset_data.append([
                asset.get("account_name", ""),
                self._format_currency(Decimal(str(asset.get("balance", 0))))
            ])
        asset_data.append(["Total Assets", self._format_currency(Decimal(str(data.get("total_assets", 0))))])
        
        elements.append(self._create_table(asset_data, highlight_last_row=True))
        elements.append(Spacer(1, 15))
        
        # Liabilities section
        elements.append(Paragraph("LIABILITIES", self.styles['SectionHeader']))
        
        liab_data = [["Account", "Amount"]]
        for liab in data.get("liabilities", []):
            liab_data.append([
                liab.get("account_name", ""),
                self._format_currency(Decimal(str(liab.get("balance", 0))))
            ])
        liab_data.append(["Total Liabilities", self._format_currency(Decimal(str(data.get("total_liabilities", 0))))])
        
        elements.append(self._create_table(liab_data, highlight_last_row=True))
        elements.append(Spacer(1, 15))
        
        # Equity section
        elements.append(Paragraph("SHAREHOLDERS' EQUITY", self.styles['SectionHeader']))
        
        equity_data = [["Account", "Amount"]]
        for eq in data.get("equity", []):
            equity_data.append([
                eq.get("account_name", ""),
                self._format_currency(Decimal(str(eq.get("balance", 0))))
            ])
        
        if include_minority_interest and data.get("minority_interest"):
            equity_data.append([
                "Non-Controlling Interest",
                self._format_currency(Decimal(str(data.get("minority_interest", 0))))
            ])
        
        equity_data.append(["Total Equity", self._format_currency(Decimal(str(data.get("total_equity", 0))))])
        
        elements.append(self._create_table(equity_data, highlight_last_row=True))
        elements.append(Spacer(1, 20))
        
        # Summary
        summary_data = [["Total Liabilities + Equity", self._format_currency(
            Decimal(str(data.get("total_liabilities", 0))) + Decimal(str(data.get("total_equity", 0)))
        )]]
        elements.append(self._create_table(summary_data, is_summary=True))
        
        # Consolidation adjustments note
        if data.get("elimination_entries"):
            elements.append(Spacer(1, 15))
            elements.append(Paragraph("Consolidation Adjustments Applied:", self.styles['SectionHeader']))
            for adj in data.get("elimination_entries", [])[:5]:
                elements.append(Paragraph(f"• {adj.get('description', '')}", self.styles['Normal']))
        
        doc.build(elements)
        content = buffer.getvalue()
        filename = f"consolidated_balance_sheet_{group_name.replace(' ', '_')}_{as_of_date.strftime('%Y%m%d')}.pdf"
        return content, filename
    
    async def _generate_consolidated_balance_sheet_excel(
        self,
        data: Dict[str, Any],
        group_name: str,
        as_of_date: date,
        include_minority_interest: bool
    ) -> Tuple[bytes, str]:
        """Generate consolidated balance sheet Excel."""
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl not available for Excel export")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated Balance Sheet"
        
        # Headers
        ws.append([group_name])
        ws.append(["Consolidated Balance Sheet"])
        ws.append([f"As of {as_of_date.strftime('%B %d, %Y')}"])
        ws.append([])
        
        # Assets
        ws.append(["ASSETS"])
        ws.append(["Account", "Amount"])
        for asset in data.get("assets", []):
            ws.append([asset.get("account_name", ""), float(asset.get("balance", 0))])
        ws.append(["Total Assets", float(data.get("total_assets", 0))])
        ws.append([])
        
        # Liabilities
        ws.append(["LIABILITIES"])
        ws.append(["Account", "Amount"])
        for liab in data.get("liabilities", []):
            ws.append([liab.get("account_name", ""), float(liab.get("balance", 0))])
        ws.append(["Total Liabilities", float(data.get("total_liabilities", 0))])
        ws.append([])
        
        # Equity
        ws.append(["SHAREHOLDERS' EQUITY"])
        ws.append(["Account", "Amount"])
        for eq in data.get("equity", []):
            ws.append([eq.get("account_name", ""), float(eq.get("balance", 0))])
        if include_minority_interest and data.get("minority_interest"):
            ws.append(["Non-Controlling Interest", float(data.get("minority_interest", 0))])
        ws.append(["Total Equity", float(data.get("total_equity", 0))])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        filename = f"consolidated_balance_sheet_{group_name.replace(' ', '_')}_{as_of_date.strftime('%Y%m%d')}.xlsx"
        return content, filename
    
    async def _generate_consolidated_balance_sheet_csv(
        self,
        data: Dict[str, Any],
        group_name: str,
        as_of_date: date
    ) -> Tuple[bytes, str]:
        """Generate consolidated balance sheet CSV."""
        import csv
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        writer.writerow([group_name])
        writer.writerow(["Consolidated Balance Sheet"])
        writer.writerow([f"As of {as_of_date.strftime('%B %d, %Y')}"])
        writer.writerow([])
        
        writer.writerow(["ASSETS"])
        writer.writerow(["Account", "Amount"])
        for asset in data.get("assets", []):
            writer.writerow([asset.get("account_name", ""), float(asset.get("balance", 0))])
        writer.writerow(["Total Assets", float(data.get("total_assets", 0))])
        writer.writerow([])
        
        writer.writerow(["LIABILITIES"])
        writer.writerow(["Account", "Amount"])
        for liab in data.get("liabilities", []):
            writer.writerow([liab.get("account_name", ""), float(liab.get("balance", 0))])
        writer.writerow(["Total Liabilities", float(data.get("total_liabilities", 0))])
        writer.writerow([])
        
        writer.writerow(["EQUITY"])
        writer.writerow(["Account", "Amount"])
        for eq in data.get("equity", []):
            writer.writerow([eq.get("account_name", ""), float(eq.get("balance", 0))])
        writer.writerow(["Total Equity", float(data.get("total_equity", 0))])
        
        content = buffer.getvalue().encode('utf-8')
        filename = f"consolidated_balance_sheet_{group_name.replace(' ', '_')}_{as_of_date.strftime('%Y%m%d')}.csv"
        return content, filename
    
    async def export_consolidated_income_statement(
        self,
        group_id: uuid.UUID,
        start_date: date,
        end_date: date,
        format: ReportFormat = ReportFormat.PDF,
    ) -> Tuple[bytes, str]:
        """
        Export consolidated income statement for an entity group.
        
        Args:
            group_id: ID of the entity group
            start_date: Period start date
            end_date: Period end date
            format: Export format (PDF, XLSX, CSV)
            
        Returns:
            Tuple of (file content bytes, filename)
        """
        from app.services.consolidation_service import ConsolidationService
        
        consol_service = ConsolidationService(self.db)
        data = await consol_service.get_consolidated_income_statement(
            group_id=group_id,
            start_date=start_date,
            end_date=end_date
        )
        
        group = await consol_service.get_entity_group(group_id)
        group_name = group.name if group else "Consolidated Group"
        
        if format == ReportFormat.PDF:
            return await self._generate_consolidated_income_statement_pdf(
                data, group_name, start_date, end_date
            )
        elif format == ReportFormat.EXCEL:
            return await self._generate_consolidated_income_statement_excel(
                data, group_name, start_date, end_date
            )
        else:  # CSV
            return await self._generate_consolidated_income_statement_csv(
                data, group_name, start_date, end_date
            )
    
    async def _generate_consolidated_income_statement_pdf(
        self,
        data: Dict[str, Any],
        group_name: str,
        start_date: date,
        end_date: date
    ) -> Tuple[bytes, str]:
        """Generate consolidated income statement PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Title
        elements.append(Paragraph(group_name, self.styles['ReportTitle']))
        elements.append(Paragraph("Consolidated Income Statement", self.styles['ReportSubtitle']))
        elements.append(Paragraph(
            f"For the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
            self.styles['ReportSubtitle']
        ))
        elements.append(Spacer(1, 20))
        
        # Revenue section
        elements.append(Paragraph("REVENUE", self.styles['SectionHeader']))
        
        rev_data = [["Account", "Amount"]]
        for rev in data.get("revenue", []):
            rev_data.append([
                rev.get("account_name", ""),
                self._format_currency(Decimal(str(rev.get("amount", 0))))
            ])
        rev_data.append(["Total Revenue", self._format_currency(Decimal(str(data.get("total_revenue", 0))))])
        
        elements.append(self._create_table(rev_data, highlight_last_row=True))
        elements.append(Spacer(1, 15))
        
        # Expenses section
        elements.append(Paragraph("EXPENSES", self.styles['SectionHeader']))
        
        exp_data = [["Account", "Amount"]]
        for exp in data.get("expenses", []):
            exp_data.append([
                exp.get("account_name", ""),
                self._format_currency(Decimal(str(exp.get("amount", 0))))
            ])
        exp_data.append(["Total Expenses", self._format_currency(Decimal(str(data.get("total_expenses", 0))))])
        
        elements.append(self._create_table(exp_data, highlight_last_row=True))
        elements.append(Spacer(1, 20))
        
        # Net Income
        net_income = data.get("net_income", 0)
        is_profit = net_income >= 0
        summary_data = [["Net Income" if is_profit else "Net Loss", self._format_currency(Decimal(str(abs(net_income))))]]
        elements.append(self._create_table(summary_data, is_summary=True, is_profit=is_profit))
        
        # Attributable income
        if data.get("attributable_to_parent") is not None:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("Attributable to:", self.styles['Normal']))
            elements.append(Paragraph(
                f"  Owners of the parent: {self._format_currency(Decimal(str(data.get('attributable_to_parent', 0))))}",
                self.styles['Normal']
            ))
            elements.append(Paragraph(
                f"  Non-controlling interests: {self._format_currency(Decimal(str(data.get('attributable_to_nci', 0))))}",
                self.styles['Normal']
            ))
        
        doc.build(elements)
        content = buffer.getvalue()
        filename = f"consolidated_income_statement_{group_name.replace(' ', '_')}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
        return content, filename
    
    async def _generate_consolidated_income_statement_excel(
        self,
        data: Dict[str, Any],
        group_name: str,
        start_date: date,
        end_date: date
    ) -> Tuple[bytes, str]:
        """Generate consolidated income statement Excel."""
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl not available for Excel export")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated P&L"
        
        # Headers
        ws.append([group_name])
        ws.append(["Consolidated Income Statement"])
        ws.append([f"For the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"])
        ws.append([])
        
        # Revenue
        ws.append(["REVENUE"])
        ws.append(["Account", "Amount"])
        for rev in data.get("revenue", []):
            ws.append([rev.get("account_name", ""), float(rev.get("amount", 0))])
        ws.append(["Total Revenue", float(data.get("total_revenue", 0))])
        ws.append([])
        
        # Expenses
        ws.append(["EXPENSES"])
        ws.append(["Account", "Amount"])
        for exp in data.get("expenses", []):
            ws.append([exp.get("account_name", ""), float(exp.get("amount", 0))])
        ws.append(["Total Expenses", float(data.get("total_expenses", 0))])
        ws.append([])
        
        # Net Income
        ws.append(["Net Income", float(data.get("net_income", 0))])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        filename = f"consolidated_income_statement_{group_name.replace(' ', '_')}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        return content, filename
    
    async def _generate_consolidated_income_statement_csv(
        self,
        data: Dict[str, Any],
        group_name: str,
        start_date: date,
        end_date: date
    ) -> Tuple[bytes, str]:
        """Generate consolidated income statement CSV."""
        import csv
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        writer.writerow([group_name])
        writer.writerow(["Consolidated Income Statement"])
        writer.writerow([f"For the period {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"])
        writer.writerow([])
        
        writer.writerow(["REVENUE"])
        writer.writerow(["Account", "Amount"])
        for rev in data.get("revenue", []):
            writer.writerow([rev.get("account_name", ""), float(rev.get("amount", 0))])
        writer.writerow(["Total Revenue", float(data.get("total_revenue", 0))])
        writer.writerow([])
        
        writer.writerow(["EXPENSES"])
        writer.writerow(["Account", "Amount"])
        for exp in data.get("expenses", []):
            writer.writerow([exp.get("account_name", ""), float(exp.get("amount", 0))])
        writer.writerow(["Total Expenses", float(data.get("total_expenses", 0))])
        writer.writerow([])
        
        writer.writerow(["Net Income", float(data.get("net_income", 0))])
        
        content = buffer.getvalue().encode('utf-8')
        filename = f"consolidated_income_statement_{group_name.replace(' ', '_')}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
        return content, filename
    
    async def export_consolidated_trial_balance(
        self,
        group_id: uuid.UUID,
        as_of_date: date,
        format: ReportFormat = ReportFormat.PDF,
    ) -> Tuple[bytes, str]:
        """
        Export consolidated trial balance for an entity group.
        
        Args:
            group_id: ID of the entity group
            as_of_date: Trial balance date
            format: Export format (PDF, XLSX, CSV)
            
        Returns:
            Tuple of (file content bytes, filename)
        """
        from app.services.consolidation_service import ConsolidationService
        
        consol_service = ConsolidationService(self.db)
        data = await consol_service.get_consolidated_trial_balance(
            group_id=group_id,
            as_of_date=as_of_date
        )
        
        group = await consol_service.get_entity_group(group_id)
        group_name = group.name if group else "Consolidated Group"
        
        if format == ReportFormat.PDF:
            return await self._generate_consolidated_trial_balance_pdf(data, group_name, as_of_date)
        elif format == ReportFormat.EXCEL:
            return await self._generate_consolidated_trial_balance_excel(data, group_name, as_of_date)
        else:  # CSV
            return await self._generate_consolidated_trial_balance_csv(data, group_name, as_of_date)
    
    async def _generate_consolidated_trial_balance_pdf(
        self,
        data: Dict[str, Any],
        group_name: str,
        as_of_date: date
    ) -> Tuple[bytes, str]:
        """Generate consolidated trial balance PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Title
        elements.append(Paragraph(group_name, self.styles['ReportTitle']))
        elements.append(Paragraph("Consolidated Trial Balance", self.styles['ReportSubtitle']))
        elements.append(Paragraph(f"As of {as_of_date.strftime('%B %d, %Y')}", self.styles['ReportSubtitle']))
        elements.append(Spacer(1, 20))
        
        # Trial balance table
        tb_data = [["Account Code", "Account Name", "Debit", "Credit"]]
        
        for account in data.get("accounts", []):
            debit = Decimal(str(account.get("debit", 0)))
            credit = Decimal(str(account.get("credit", 0)))
            tb_data.append([
                account.get("account_code", ""),
                account.get("account_name", ""),
                self._format_currency(debit) if debit > 0 else "",
                self._format_currency(credit) if credit > 0 else ""
            ])
        
        # Totals
        total_debit = Decimal(str(data.get("total_debit", 0)))
        total_credit = Decimal(str(data.get("total_credit", 0)))
        tb_data.append(["", "TOTALS", self._format_currency(total_debit), self._format_currency(total_credit)])
        
        table = Table(tb_data, colWidths=[1*inch, 3*inch, 1.25*inch, 1.25*inch])
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor("#cbd5e0")),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor("#2d3748")),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        # Balance check
        elements.append(Spacer(1, 15))
        is_balanced = abs(total_debit - total_credit) < Decimal("0.01")
        if is_balanced:
            elements.append(Paragraph("✓ Trial balance is in balance", self.styles['Normal']))
        else:
            elements.append(Paragraph(
                f"⚠ Out of balance by: {self._format_currency(abs(total_debit - total_credit))}",
                self.styles['Normal']
            ))
        
        doc.build(elements)
        content = buffer.getvalue()
        filename = f"consolidated_trial_balance_{group_name.replace(' ', '_')}_{as_of_date.strftime('%Y%m%d')}.pdf"
        return content, filename
    
    async def _generate_consolidated_trial_balance_excel(
        self,
        data: Dict[str, Any],
        group_name: str,
        as_of_date: date
    ) -> Tuple[bytes, str]:
        """Generate consolidated trial balance Excel."""
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl not available for Excel export")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated TB"
        
        ws.append([group_name])
        ws.append(["Consolidated Trial Balance"])
        ws.append([f"As of {as_of_date.strftime('%B %d, %Y')}"])
        ws.append([])
        
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        for account in data.get("accounts", []):
            ws.append([
                account.get("account_code", ""),
                account.get("account_name", ""),
                float(account.get("debit", 0)),
                float(account.get("credit", 0))
            ])
        
        ws.append(["", "TOTALS", float(data.get("total_debit", 0)), float(data.get("total_credit", 0))])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        filename = f"consolidated_trial_balance_{group_name.replace(' ', '_')}_{as_of_date.strftime('%Y%m%d')}.xlsx"
        return content, filename
    
    async def _generate_consolidated_trial_balance_csv(
        self,
        data: Dict[str, Any],
        group_name: str,
        as_of_date: date
    ) -> Tuple[bytes, str]:
        """Generate consolidated trial balance CSV."""
        import csv
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        writer.writerow([group_name])
        writer.writerow(["Consolidated Trial Balance"])
        writer.writerow([f"As of {as_of_date.strftime('%B %d, %Y')}"])
        writer.writerow([])
        
        writer.writerow(["Account Code", "Account Name", "Debit", "Credit"])
        for account in data.get("accounts", []):
            writer.writerow([
                account.get("account_code", ""),
                account.get("account_name", ""),
                float(account.get("debit", 0)),
                float(account.get("credit", 0))
            ])
        
        writer.writerow(["", "TOTALS", float(data.get("total_debit", 0)), float(data.get("total_credit", 0))])
        
        content = buffer.getvalue().encode('utf-8')
        filename = f"consolidated_trial_balance_{group_name.replace(' ', '_')}_{as_of_date.strftime('%Y%m%d')}.csv"
        return content, filename